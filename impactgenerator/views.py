from datetime import datetime
from django import forms
from django.forms.widgets import DateTimeInput
from django.shortcuts import render, redirect
from django.core.files.storage import FileSystemStorage
from django.core.files import File
from django.core.files.base import ContentFile, File
from django.http import HttpResponse
from .forms import DocumentForm
from .models import Document
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.conf import settings
import datetime

# # def home(request):
# #     if request.method == "POST":
# #         document_name = request.POST['name']
# #         csv = request.FILES['document']
# #         read_excel = pd.read_csv(request.FILES['document'])
# #         document = Document.objects.create(
# #             name=document_name, csv=csv, read_excel=csv, write_excel=csv)
# #         document.refresh_from_db()

# #         id = document.id
# #         print(id)
# #         file = document.csv
# #         print(file)

# #         read_file = pd.read_csv(file)

# #         response = HttpResponse(content_type='text/xlsx')

# #         print(response)
# #         read_file.to_excel(response)

# #         read_wb = openpyxl.load_workbook("Impact analysis.xlsx")

# #         # sheets = read_wb.sheetnames
# #         # print(sheets)
# #         request.FILES['document'.to_excel(r'Imp')]

# #         # To create a new workbook
# #         write_wb = Workbook()
# #         write_wb['Sheet'].title = "Impact Analysis"
# #         write_sheet = write_wb.active

# #         write_sheet['A1'] = "S.No"
# #         write_sheet['B1'] = "Status"
# #         write_sheet['C1'] = "Bug ID"
# #         write_sheet['D1'] = "Bug Description"
# #         write_sheet['E1'] = "DDC"
# #         write_sheet['F1'] = "Area of Fix from GE(file name)"
# #         write_sheet['G1'] = "Before Fix"
# #         write_sheet['H1'] = "After Fix"
# #         write_sheet['I1'] = "Scenario description in detail"
# #         write_sheet['J1'] = "Expected output"
# #         write_sheet['K1'] = "CH LL review status (For J ~ N col)"
# #         write_sheet['L1'] = "Retest Result"
# #         write_sheet['M1'] = "Executed on (DD/MM/YYYY)"
# #         write_sheet['N1'] = "CH Bug ID"
# #         write_sheet['O1'] = "Comments"

# #         # to get the current sheet name
# #         sheet_name = read_wb.active.title
# #         # print(sheet_name)

# #         # reading the contents of the current sheet
# #         sheet1 = read_wb[sheet_name]
# #         # print(sheet1)

# #         # to know the maximum countber of columns
# #         column_end = sheet1.max_column

# #         row = 1

# #         for col in range(1, column_end+1):
# #             # print(row, col)
# #             column_header = sheet1.cell(row, col).value
# #             print(column_header)
# #             # print(len(column_header))
# #             if column_header == "Status":
# #                 print("found")
# #                 col_value = col
# #                 # print(col_value)
# #                 count = 2
# #                 for row_count in range(2, sheet1.max_row):
# #                     result = sheet1.cell(row_count, col_value).value

# #                     write_sheet['B'+str(count)] = result
# #                     count = count + 1
# #                     # print(count)
# #             elif column_header == "#":
# #                 print("found")
# #                 col_value = col
# #                 # print(col_value)
# #                 count = 2
# #                 for row_count in range(2, sheet1.max_row):
# #                     result = sheet1.cell(row_count, col_value).value

# #                     write_sheet['C'+str(count)] = result
# #                     count = count + 1
# #                     # print(count)
# #             elif column_header == "Description":
# #                 print("found")
# #                 col_value = col
# #                 # print(col_value)
# #                 count = 2
# #                 for row_count in range(2, sheet1.max_row):
# #                     result = sheet1.cell(row_count, col_value).value
# #                     # print(result)
# #                     result = result.lower()

# #                     desc_start = result.index("steps")
# #                     desc_string = result[desc_start:]
# #                     desc_end = desc_string.index("\n\n")
# #                     description = desc_string[:desc_end]
# #                     print(description)
# #                     write_sheet['I'+str(count)] = description.strip()
# #                     write_sheet['I'+str(count)
# #                                 ].alignment = Alignment(wrapText=True)
# #                     # print(result.index("expected"))

# #                     expected_result_start = result.index("expected")
# #                     expected_string = result[expected_result_start:]
# #                     expected_result_end = expected_string.index('\n\n')
# #                     expected_result = expected_string[:expected_result_end]
# #                     print(expected_result)
# #                     write_sheet['H'+str(count)] = expected_result.strip()
# #                     write_sheet['H'+str(count)
# #                                 ].alignment = Alignment(wrapText=True)

# #                     write_sheet['J'+str(count)] = expected_result.strip()
# #                     write_sheet['J'+str(count)
# #                                 ].alignment = Alignment(wrapText=True)

# #                     actual_result_start = result.index("actual")
# #                     actual_string = result[actual_result_start:]
# #                     actual_result_end = actual_string.index('\n\n')
# #                     actual_result = actual_string[:actual_result_end]
# #                     print(actual_result)
# #                     write_sheet['G'+str(count)] = actual_result.strip()
# #                     write_sheet['G'+str(count)
# #                                 ].alignment = Alignment(wrapText=True)

# #                     count = count + 1
# #                     # print(count)

# #             elif column_header == "Details Defect Category":
# #                 print("found")
# #                 col_value = col
# #                 # print(col_value)
# #                 count = 2
# #                 for row_count in range(2, sheet1.max_row):
# #                     result = sheet1.cell(row_count, col_value).value

# #                     write_sheet['E'+str(count)] = result.strip()
# #                     count = count + 1

# #             elif column_header == "Subject":
# #                 print("found")
# #                 col_value = col
# #                 # print(col_value)
# #                 count = 2
# #                 for row_count in range(2, sheet1.max_row):
# #                     result = sheet1.cell(row_count, col_value).value

# #                     write_sheet['D'+str(count)] = result.strip()
# #                     count = count + 1

# #             write_wb.save("Impact Analysis Document.xlsx")

# #             # if form.is_valid():
# #             #     form.save()
# #             return redirect('home')
# #     else:
# #         # form = DocumentForm()
# #         # context = {
# #         #     'form': form,
# #         return render(request, 'impactgenerator/home.html', context={})


# # # def upload(request):
# # #     if request.method == "POST":
# # #         uploaded_file = request.FILES['document']
# # #         fs = FileSystemStorage()
# # #         fs.save(uploaded_file.name, uploaded_file)
# # #         print(uploaded_file.name)
# # #         print(uploaded_file.size)
# # #     return render(request, 'impactgenerator/upload.html', context={})


# # # def document(request):
# # #     return render(request, 'impactgenerator/document.html', context={})


# # # def upload_document(request):

# # #     if request.method == "POST":
# # #         form = DocumentForm(request.POST, request.FILES)
# # #         if form.is_valid():
# # #             form.save()
# # #             return redirect('document')
# # #     else:
# # #         form = DocumentForm()
# # #         context = {
# # #             'form': form,
# # #         }

# # #     return render(request, 'impactgenerator/upload_document.html', context)


# def home(request):
#     if request.method == "POST":
#         document_name = request.POST['name']
#         csv = request.FILES['document']

#         document = Document.objects.create(
#             name=document_name, csv=csv, read_excel=csv, write_excel=csv)
#         document.refresh_from_db()

#         id = document.id
#         print(id)
#         file = document.csv
#         print(file)

#         read_file = pd.read_csv(file)

#         response = HttpResponse(content_type="application/ms-excel")
#         response['Content-Dispositon'] = 'attachment; filename="ImpactAnalysis_read"' + \
#             str(datetime.datetime.now())+'.xls'

#         rb = read_file.to_excel()
#         print(rb)
#     return render(request, 'impactgenerator/home.html', context={})


# def export_excel(request):
#     pass


def home(request):
    pass
